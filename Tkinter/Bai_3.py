import cv2 as cv
import tkinter as tk
import keras
import os
from tkinter import *
from tkinter import messagebox, filedialog as fd

from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws['A1'] = "Tên File"
ws['B1'] = "Phân loại"

window = Tk()
window.title("Welcome to my project")
model = keras.models.load_model('train_model.pb')

def open_1():
    global img_1,img_2
    filename = fd.askdirectory()
    def prepare(filepath):
        img_size = 244
        img_array = cv.imread(filepath, cv.IMREAD_GRAYSCALE)
        new_array = cv.resize(img_array, (img_size, img_size))
        new_array = new_array / 255.0
        return new_array.reshape(-1, img_size, img_size, 1)
    save_path = "D:/Thực tập tốt nghiệp/Tuần 3/Bai_3.xlsx"
    for img in os.listdir(filename):
        path_img = os.path.join(filename, img)
        basename = os.path.basename(img)
        name = os.path.splitext(basename)[0]
        prediction = model.predict([prepare('{}'.format(path_img))])
        a = prediction[0]
        c = 0
        if a[0] >= 0.9:
            b = 'Cat'
            c = a[0]
        elif a[1] >= 0.9:
            b = 'Dog'
            c = a[1]
        elif a[2] >= 0.9:
            b = 'Panda'
            c = a[2]
        else:
            b = 'unknown'
        ws.append([name, b])
        wb.save(save_path)
    messagebox.showwarning('Message', 'File kết quả được lưu tại {}'.format(save_path))

label = tk.Label(window, text = "Chào bạn đã đến đây", font = ("Arial bold", 16)).pack()
label_2 = tk.Label(window, text = "Bấm nút 'Chose Folder'", font = ("Arial bold", 13)).pack()
label_3 = tk.Label(window, text = "để chọn folder hình ảnh", font = ("Arial bold", 13)).pack()
label_4 = tk.Label(window, text = "cần phân loại", font = ("Arial bold", 13)).pack()
chose_button = tk.Button(window, text="Chose Folder",command=open_1).pack()
window.geometry("240x180")
window.mainloop()


