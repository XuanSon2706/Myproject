import cv2 as cv
import numpy as np
import keras
import os

from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws['A1'] = "TenAnh"
ws['B1'] = "Phanloai"

categories = ["glioma_tumor","meningioma_tumor","no_tumor","pituitary_tumor"]
categories_1 = ["pituitary_tumor"]
datadir = ("Training")

img_size = 244
def prepare(filepath):
    img_array = cv.imread(filepath,cv.IMREAD_GRAYSCALE)
    new_array = cv.resize(img_array, (img_size, img_size))
    new_array = new_array / 255.0
    return new_array.reshape(-1, img_size, img_size, 1)

model = keras.models.load_model('train_model.model')

for category in categories_1:
    path = os.path.join(datadir,category)
    for img in os.listdir(path):
        basename = os.path.basename(img)
        name = os.path.splitext(basename)[0]
        path_img = os.path.join(path, img)
        prediction = model.predict([prepare('{}'.format(path_img))])
        print(prediction)
        dudoan = categories[int(np.argmax(prediction))]
        ws.append([name,dudoan])
        wb.save("D:/Thực tập tốt nghiệp/Tuần 2/test2/test2.xlsx")
