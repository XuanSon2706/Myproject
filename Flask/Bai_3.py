from flask import Flask, request, render_template
import os
from werkzeug.utils import secure_filename
import cv2 as cv
import keras
from openpyxl import Workbook
import pandas as pd

wb = Workbook()
ws = wb.active
ws['A1'] = "Tên File"
ws['B1'] = "Phân loại"
save_path = "static/Bai_3.xlsx"

model = keras.models.load_model('train_model.pb')

app = Flask(__name__)

UPLOAD_FOLDER = 'static/uploads_1/'

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

def prepare(filepath):
    img_size = 244
    img_array = cv.imread(filepath, cv.IMREAD_GRAYSCALE)
    new_array = cv.resize(img_array, (img_size, img_size))
    new_array = new_array / 255.0
    return new_array.reshape(-1, img_size, img_size, 1)

def predict(path_img):
    prediction = model.predict([prepare('{}'.format(path_img))])
    a = prediction[0]
    c = 0
    if a[0] >= 0.9:
        b = 'Mèo'
        c = a[0]
    elif a[1] >= 0.9:
        b = 'Chó'
        c = a[1]
    elif a[2] >= 0.9:
        b = 'Gấu trúc'
        c = a[2]
    else:
        b = 'Unknow'
        c = a[3]
    return b,c

@app.route('/')
def home():
    return render_template('index_1.html')


@app.route('/', methods=['GET','POST'])
def upload_image():
    files = request.files.getlist("file")
    for file in files:
        filename = secure_filename(file.filename)
        path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(path)
        b, c = predict(path)
        basename = os.path.basename(filename)
        name = os.path.splitext(basename)[0]
        ws.append([name, b])
        wb.save(save_path)
        data = pd.read_excel(save_path)
    return render_template('index_2.html',data = data.to_html())

if __name__ == "__main__":
    app.run()
